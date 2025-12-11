"""Module để load và sử dụng mapping từ file xlsx."""
import os
from pathlib import Path
from typing import Dict, Any, Optional
from .utils import normalize_text
import json
# Cache cho mappings
_mappings_cache: Dict[str, Dict[str, Any]] = {}
     # thư mục của file .py


def _load_mappings():
    """Load tất cả mappings từ file xlsx và cache lại."""
    global _mappings_cache
    
    if _mappings_cache:
        return _mappings_cache
    
    try:
        from openpyxl import load_workbook
        from pathlib import Path
        
        # Tìm file xlsx
        project_root = Path(__file__).resolve().parents[1]
        xlsx_path = project_root / "output" / "map.xlsx"
        
        if not xlsx_path.exists():
            print(f"[Mapping] File {xlsx_path} không tồn tại")
            return {}
        
        wb = load_workbook(xlsx_path, data_only=True)

        def get_safe_int(v):
            if v is None:
                return None
            try:
                if isinstance(v, float):
                    return int(v) if v.is_integer() else v
                if isinstance(v, str) and v.replace('.', '').isdigit():
                    f = float(v)
                    return int(f) if f.is_integer() else f
                return int(v)
            except:
                return None
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            mapping = {}
            
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            
            # Tìm header row
            header_row_idx = None
            for idx, row in enumerate(rows):
                if row and any(cell and str(cell).upper() in ["ID", "VALUE", "SLUG"]
                               for cell in row):
                    header_row_idx = idx
                    break
            
            if header_row_idx is None:
                continue
            
            header_row = rows[header_row_idx]
            col_mapping = {}
            for col_idx, cell in enumerate(header_row):
                if cell is not None:
                    col_mapping[str(cell).upper().strip()] = col_idx
            
            if "ID" not in col_mapping or "VALUE" not in col_mapping:
                print(f"[Mapping] Sheet '{sheet_name}' thiếu cột ID hoặc VALUE, bỏ qua.")
                continue
            
            id_col_idx = col_mapping["ID"]
            value_col_idx = col_mapping["VALUE"]
            slug_col_idx = col_mapping.get("SLUG")

            print(f"[Mapping] Sheet '{sheet_name}': ID@{id_col_idx+1}, "
                  f"Value@{value_col_idx+1}, Slug@{slug_col_idx+1 if slug_col_idx else 'N/A'}")
            
            # Parse data rows
            for row in rows[header_row_idx + 1:]:
                if not row or not any(row):
                    continue
                
                try:
                    # ID
                    id_cell = row[id_col_idx] if id_col_idx < len(row) else None
                    if id_cell is None:
                        continue
                    id_val = get_safe_int(id_cell)
                    
                    # VALUE
                    value_cell = row[value_col_idx] if value_col_idx < len(row) else None
                    value = str(value_cell).strip() if value_cell else None
                    if not value:
                        continue
                    
                    # SLUG optional
                    slug = None
                    if slug_col_idx is not None and slug_col_idx < len(row):
                        slug_cell = row[slug_col_idx]
                        slug = str(slug_cell).strip() if slug_cell else None
                    
                    # --------- TẠO ENTRY MỞ RỘNG -----------
                    entry = {"id": id_val}

                    # district_id → cần province_id
                    if sheet_name.lower() == "district_id":
                        if "PROVINCE_ID" in col_mapping:
                            pidx = col_mapping["PROVINCE_ID"]
                            entry["province_id"] = get_safe_int(row[pidx]) if pidx < len(row) else None

                    # ward_id → cần district_id + province_id
                    if sheet_name.lower() == "ward_id":
                        if "DISTRICT_ID" in col_mapping:
                            didx = col_mapping["DISTRICT_ID"]
                            entry["district_id"] = get_safe_int(row[didx]) if didx < len(row) else None
                        if "PROVINCE_ID" in col_mapping:
                            pidx = col_mapping["PROVINCE_ID"]
                            entry["province_id"] = get_safe_int(row[pidx]) if pidx < len(row) else None
                    
                    # --------- LƯU MAPPING THEO NHIỀU KEY -----------
                    def store_key(k):
                        if k:
                            mapping[k] = entry
                    
                    value_lower = value.lower()
                    normalized_value = normalize_text(value)
                    
                    store_key(value_lower)
                    store_key(normalized_value)

                    if slug:
                        slug_lower = slug.lower()
                        normalized_slug = normalize_text(slug)
                        store_key(slug_lower)
                        store_key(normalized_slug)
                
                except Exception as e:
                    print(f"[Mapping] Lỗi parse row trong sheet '{sheet_name}': {e}")
                    continue
            
            if mapping:
                _mappings_cache[sheet_name] = mapping
                print(f"[Mapping] Loaded {len(mapping)} entries from sheet '{sheet_name}'")
        
        wb.close()
        return _mappings_cache
    
    except ImportError:
        print("[Mapping] openpyxl chưa được cài đặt, không thể load mappings")
        return {}
    except Exception as e:
        print(f"[Mapping] Lỗi khi load mappings: {e}")
        return {}



def find_ward_key_loose(json_file = "", name = "", province_id=None, district_id=None):
    project_root = Path(__file__).resolve().parents[1]
    json_path = project_root / "output" / json_file
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    name_norm = name.strip().lower()

    for key, value in data.items():
        if district_id:
            if (
                value.get("province_id") == province_id
                and float(value.get("district_id")) == float(district_id)
                and name_norm in value.get("name", "").lower()
            ):
                return key
        else:
            if (
                value.get("province_id") == province_id
                and name_norm in value.get("name", "").lower()
            ):
                return key            

    return None

def partial_match(value_lower: str, key: str) -> bool:
    """Word-based partial match: Tất cả từ của value có trong key (split by space/-), hoặc ngược lại."""
    value_words = set(value_lower.split())
    key_words = set(key.replace('-', ' ').lower().split())
    return value_words.issubset(key_words) or key_words.issubset(value_words)

def get_mapping(sheet_name: str, value: str, filter_slug_parts: Optional[list[str]] = None, return_entry: bool = False) -> Optional[Any]:
    """
    Nâng cấp: Cho phép match ward/district/province theo kiểu chứa (contains),
    không cần có từ 'phường/xã/thị trấn'.
    """

    mappings = _load_mappings()
    sheet_mapping = mappings.get(sheet_name) or mappings.get(sheet_name.strip(), {})
    if not sheet_mapping or not value:
        return None

    value_norm = normalize_text(value)

    # ---------------------------------
    # 1. EXACT MATCH
    # ---------------------------------
    entry = sheet_mapping.get(value_norm)
    if entry:
        return entry["id"]

    # Chuẩn bị filter context
    filter_words = set()
    if filter_slug_parts:
        for part in filter_slug_parts:
            if part:
                filter_words.update(normalize_text(part).split())

    # ---------------------------------
    # 2. PARTIAL MATCH (word-based)
    # ---------------------------------
    for key, entry in sheet_mapping.items():
        key_norm = key
        key_words = set(key_norm.split())

        if not partial_match(value_norm, key_norm):
            continue

        if filter_words and not filter_words.issubset(key_words):
            continue

        if sheet_name.lower() == "district_id" and filter_slug_parts:
            prov_name = filter_slug_parts[0]
            prov_id = get_mapping("province_id", prov_name)
            if prov_id and entry.get("province_id") and entry["province_id"] != prov_id:
                continue

        if sheet_name.lower() == "ward_id" and filter_slug_parts:
            prov_name = filter_slug_parts[0] if len(filter_slug_parts) > 0 else None
            dist_name = filter_slug_parts[1] if len(filter_slug_parts) > 1 else None

            if prov_name:
                prov_id = get_mapping("province_id", prov_name)
                if prov_id and entry["province_id"] != prov_id:
                    continue

            if dist_name:
                dist_id = get_mapping("district_id", dist_name)
                if dist_id and entry["district_id"] != dist_id:
                    continue

        return entry["id"]

    # ---------------------------------
    # 3. CONTAINS MATCH (NEW)
    #    Với ward, district, province:
    #    Nếu value nằm trong key hoặc key nằm trong value → cho phép match
    # ---------------------------------
    for key, entry in sheet_mapping.items():
        key_norm = key

        if value_norm in key_norm or key_norm in value_norm:
            # Context filter
            if sheet_name.lower() == "ward_id" and filter_slug_parts:
                prov_name = filter_slug_parts[0] if len(filter_slug_parts) > 0 else None
                dist_name = filter_slug_parts[1] if len(filter_slug_parts) > 1 else None

                if prov_name:
                    prov_id = get_mapping("province_id", prov_name)
                    if prov_id and entry["province_id"] != prov_id:
                        continue

                if dist_name:
                    dist_id = get_mapping("district_id", dist_name)
                    if dist_id and entry["district_id"] != dist_id:
                        continue

            return entry if return_entry else entry["id"]

    return None




def get_all_mappings() -> Dict[str, Dict[str, Any]]:
    """Lấy tất cả mappings đã load."""
    return _load_mappings()

